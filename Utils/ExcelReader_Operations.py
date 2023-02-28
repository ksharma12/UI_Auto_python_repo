"""
Basic info about excel_sheet via openpyxl module:
1. Excel files are called Workbooks.
2. Each Workbook can contain multiple sheets.
3. Every sheet consists of rows starting from 1 and columns starting from A.
4. Rows and columns together make up a cell.
5. Any type of data can be stored.
6. pip install openpyxl - for installing openpyxl module
"""
import traceback
from openpyxl import load_workbook


class ExcelReader_Helpers:

    # Provide the file location for the Excel file you want to open in Python.
    def __init__(self, file_path, sheet_name):
        try:
            self.file_path = file_path
            workbook = load_workbook(self.file_path)
            self.sheet = workbook[sheet_name]
        except:
            print(traceback.print_exc())

    # Read single using cell_name can be [A2,H16,J4,D5...]
    def read_cell_data_via_cellName(self, cell_name):
        global cell_data
        try:
            cell_data = self.sheet[cell_name].value
        except:
            print(traceback.print_exc())
        return cell_data

    # Read single cell data using row_no/colum_no = integers[1,2,3,4,5....]
    def read_cell_data_via_rowNo_columnNo(self, row_no, column_no):
        global cell_data
        try:
            cell_data = self.sheet.cell(row=row_no, column=column_no).value
        except:
            print(traceback.print_exc())
        return cell_data

    # Reading a range of cells using cell names A1,H18,G12....
    # Return dict with dic[row]: 1,2,3,4 and corresponding values
    def read_all_data_in_range(self, start_cell, end_cell):
        complete_data_dict = {}
        row_no = 1
        try:
            for row in self.sheet[f"{start_cell}:{end_cell}"]:
                complete_data_dict[row_no] = [x.value for x in row]
                row_no += 1
        except:
            print(traceback.print_exc())
        return complete_data_dict

    # Access all cells in row i [i = 1,2,3,4,5...]
    # Return list of respective row
    def read_single_row_data(self, rowNo):
        row_lst = []
        try:
            row_lst = [data.value for data in self.sheet[f"{str(rowNo)}"] if data.value is not None]
        except:
            print(traceback.print_exc())
        return row_lst

    # Read complete data in Excel sheet
    # Return dict with dic[row]: 1,2,3,4 and corresponding values
    def read_all_rows_data(self):
        complete_row_data_dict = {}
        row_no = 1
        try:
            for row in self.sheet.rows:
                lst = [data.value for data in row if data.value is not None]
                if len(lst) == 0:
                    break
                complete_row_data_dict[row_no] = lst
                row_no += 1
        except:
            print(traceback.print_exc())
        return complete_row_data_dict

    # Read column data using column name [A,B,C,D,E.....]
    # Return list of column data can be accessed using
    def read_column_via_cell_name(self, columnName):
        column_lst = []
        try:
            # Access all cells in column A
            column_lst = [data.value for data in self.sheet[f"{columnName}"] if data.value is not None]
        except:
            print(traceback.print_exc())
        return column_lst

    # Read complete data in Excel sheet
    # Return dict with dic[column]: 1,2,3,4 and corresponding values
    def read_all_columns_data(self):
        global complete_column_data_dict
        column_no = 1
        try:
            for column in self.sheet.columns:
                lst = [data.value for data in column if data.value is not None]
                if len(lst) == 0:
                    break
                complete_column_data_dict[column_no] = lst
                column_no += 1
        except:
            print(traceback.print_exc())
        return complete_column_data_dict

    # Return total rows in Excel sheet
    def total_rows_count(self):
        global count
        try:
            count = 0
            all_row_data_dic = ExcelReader_Helpers.read_all_rows_data(self)
            for row in all_row_data_dic.values():
                count += 1
                if len(row) == 0:
                    break
        except:
            print(traceback.print_exc())
        return count

    # Return total columns in Excel sheet
    def total_columns_count(self):
        global count
        try:
            count = 0
            all_row_data_dic = ExcelReader_Helpers.read_all_columns_data(self)
            for column in all_row_data_dic.values():
                count += 1
                if len(column) == 0:
                    break
        except:
            print(traceback.print_exc())
        return count

    # Access all cells from between
    # "Row 1 and Row 2" and "Column 1 and Column 3"
    # Return dict[i] // i = rowNo. 1,2,3,4,5,6....
    def read_row_data_between_min_max_rows_columns(self, min_row, max_row, min_col, max_col):
        global complete_row_data_dict
        try:
            complete_row_data_dict = {}
            no = 1
            for row in self.sheet.iter_rows(min_row=min_row,
                                            max_row=max_row,
                                            min_col=min_col,
                                            max_col=max_col):
                lst = [data.value for data in row if data.value is not None]
                if len(lst) == 0:
                    break
                complete_row_data_dict[no] = lst
                no += 1
        except:
            print(traceback.print_exc())
        return complete_row_data_dict

    # Access all cells from between
    # "Row 1 and Row 2"
    # Return dict[i] // i = rowNo. 1,2,3,4,5,6....
    def read_row_data_between_min_max_rows(self, min_row, max_row):
        global complete_row_data_dict
        try:
            complete_row_data_dict = {}
            no = 1
            for row in self.sheet.iter_rows(min_row=min_row,
                                            max_row=max_row):
                lst = [data.value for data in row if data.value is not None]
                if len(lst) == 0:
                    break
                complete_row_data_dict[no] = lst
                no += 1
        except:
            print(traceback.print_exc())
        return complete_row_data_dict

    # Access all cells from between
    # "Row 1 and Row 2" and "Column 1 and Column 3"
    # Return dict[i] // i = columnNo. 1,2,3,4,5,6....
    def read_column_data_between_min_max_rows_columns(self, min_row, max_row, min_col, max_col):
        global complete_column_data_dict
        try:
            complete_column_data_dict = {}
            no = 1
            for column in self.sheet.iter_cols(min_row=min_row,
                                               max_row=max_row,
                                               min_col=min_col,
                                               max_col=max_col):
                lst = [data.value for data in column if data.value is not None]
                if len(lst) == 0:
                    break
                complete_column_data_dict[no] = lst
                no += 1
        except:
            print(traceback.print_exc())
        return complete_column_data_dict

    # Access all columns data
    # Return dict[i] // i = columnNo. 1,2,3,4,5,6....
    def read_all_column_via_iter_cols(self):
        global complete_column_data_dict
        try:
            complete_column_data_dict = {}
            no = 1
            for column in self.sheet.iter_cols():
                lst = [data.value for data in column if data.value is not None]
                if len(lst) == 0:
                    break
                complete_column_data_dict[no] = lst
                no += 1
        except:
            print(traceback.print_exc())
        return complete_column_data_dict

    # Access all columns data
    # Return dict[i] // i = columnNo. 1,2,3,4,5,6....
    def read_all_rows_via_iter_rows(self):
        global complete_row_data_dict
        try:
            complete_row_data_dict = {}
            no = 1
            for row in self.sheet.iter_rows():
                lst = [data.value for data in row if data.value is not None]
                if len(lst) == 0:
                    break
                complete_row_data_dict[no] = lst
                no += 1
        except:
            print(traceback.print_exc())
        return complete_row_data_dict


filepath = '/Automation_Framework/DataSet/speadsheet1.xlsx'
obj = ExcelReader_Helpers(file_path=filepath, sheet_name="ClientInfo")
print(obj.read_column_data_between_min_max_rows_columns(1, 76, 1, 16))
